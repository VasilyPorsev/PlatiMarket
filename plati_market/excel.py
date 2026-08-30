from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

from .scraper import Offer


def _fill_sheet(sheet, offers: list[Offer]) -> None:
    sheet.append(["Ссылка", "Продавец", "Название тарифа", "Продано", "Возвратов", "Цена"])
    for offer in offers:
        sheet.append([offer.url, offer.seller, offer.tariff, offer.sold, offer.returns, float(offer.price)])
        link = sheet.cell(sheet.max_row, 1)
        link.hyperlink = offer.url
        link.style = "Hyperlink"

    for cell in sheet[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="4472C4")
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = f"A1:F{max(sheet.max_row, 1)}"
    sheet.column_dimensions["A"].width = 75
    sheet.column_dimensions["B"].width = 28
    sheet.column_dimensions["C"].width = 80
    sheet.column_dimensions["D"].width = 14
    sheet.column_dimensions["E"].width = 14
    sheet.column_dimensions["F"].width = 15
    for cell in sheet["F"][1:]:
        cell.number_format = '#,##0.00 "₽"'
def save_excel(offers: list[Offer], output: str | Path, pro_offers: list[Offer] | None = None) -> Path:
    path = Path(output).resolve()
    path.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    plus_sheet = workbook.active
    plus_sheet.title = "ChatGPT Plus"
    _fill_sheet(plus_sheet, offers)
    pro_sheet = workbook.create_sheet("ChatGPT Pro")
    _fill_sheet(pro_sheet, pro_offers or [])
    workbook.save(path)
    return path
